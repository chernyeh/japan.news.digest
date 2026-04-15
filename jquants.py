"""
jquants.py — J-Quants API integration for Japan Investment Digest

Handles:
  - Authentication (refresh token → ID token flow)
  - Earnings Calendar (/v2/equities/earnings-calendar)
  - Financial Summary (/v2/fins/summary)

Free plan limitations:
  - Stock prices / financial data: 2 years history, 12-week delayed
  - Earnings calendar: available, covers March/September fiscal year-end companies only
  - No CSV bulk download (API access only)

Usage:
  api_key = get_jquants_secret()
  token   = get_id_token(api_key)
  cal     = fetch_earnings_calendar(token)
  summary = fetch_financial_summary(token, code="7203")
"""

import requests
import os
from datetime import datetime, timedelta, date
from functools import lru_cache
import time


# ── Secret retrieval ──────────────────────────────────────────────────────────

def get_jquants_secret() -> str:
    """Get J-Quants API key from Streamlit secrets or environment."""
    try:
        import streamlit as st
        val = st.secrets.get("JQUANTS_API_KEY", "")
        if val:
            return val
    except Exception:
        pass
    return os.environ.get("JQUANTS_API_KEY", "")


# ── Authentication ────────────────────────────────────────────────────────────
# J-Quants uses a two-step auth:
#   1. POST /v1/token/auth_user with email+password → refresh token (valid 1 week)
#   2. POST /v1/token/auth_refresh with refresh token → ID token (valid 24h)
# But with API key auth (the recommended method for free tier), you can use the
# API key directly as the x-api-key header on all requests.

BASE = "https://api.jquants.com/v2"
HEADERS_TEMPLATE = {"Content-Type": "application/json"}


def _headers(api_key: str) -> dict:
    return {"x-api-key": api_key, "Content-Type": "application/json"}


def test_connection(api_key: str) -> tuple:
    """
    Test that the API key works.
    Returns (True, "OK") or (False, "error message").
    """
    if not api_key:
        return False, "No JQUANTS_API_KEY found in Streamlit Secrets."
    try:
        resp = requests.get(
            f"{BASE}/markets/calendar",
            headers=_headers(api_key),
            timeout=10,
        )
        if resp.status_code == 200:
            return True, "Connected"
        elif resp.status_code == 401:
            return False, "Invalid API key — check JQUANTS_API_KEY in Streamlit Secrets."
        elif resp.status_code == 403:
            return False, "Access denied — your plan may not include this endpoint."
        else:
            return False, f"HTTP {resp.status_code}: {resp.text[:100]}"
    except Exception as e:
        return False, f"Connection error: {e}"


# ── Earnings Calendar ─────────────────────────────────────────────────────────

def fetch_earnings_calendar(api_key: str) -> list:
    """
    Fetch the earnings announcement schedule.
    Returns list of dicts: {Date, Code, CoName, FY, SectorNm, FQ, Section}

    Note: Only covers March and September fiscal year-end companies for now.
    Date field is empty string if announcement date is not yet confirmed.
    """
    if not api_key:
        return []

    results = []
    pagination_key = None

    while True:
        params = {}
        if pagination_key:
            params["pagination_key"] = pagination_key

        try:
            resp = requests.get(
                f"{BASE}/equities/earnings-calendar",
                headers=_headers(api_key),
                params=params,
                timeout=20,
            )
            if resp.status_code != 200:
                print(f"Earnings calendar error: HTTP {resp.status_code} — {resp.text[:100]}")
                break

            data = resp.json()
            batch = data.get("data", [])
            results.extend(batch)

            pagination_key = data.get("pagination_key")
            if not pagination_key or not batch:
                break

        except Exception as e:
            print(f"Earnings calendar fetch error: {e}")
            break

    return results


def group_calendar_by_date(entries: list) -> dict:
    """
    Group earnings calendar entries by reporting date.
    Returns dict: {date_str: [entries]} sorted by date.
    Entries with no confirmed date go under key "TBD".
    """
    grouped = {}
    for entry in entries:
        d = entry.get("Date", "")
        key = d if d else "TBD"
        grouped.setdefault(key, []).append(entry)

    # Sort: dated entries first (ascending), TBD at end
    sorted_keys = sorted([k for k in grouped if k != "TBD"]) + (["TBD"] if "TBD" in grouped else [])
    return {k: grouped[k] for k in sorted_keys}


def label_date_bucket(date_str: str) -> str:
    """
    Label a date string relative to today:
    'Today', 'Tomorrow', 'This Week', 'Next Week', 'Next 30 Days', 'Later', 'TBD'
    """
    if not date_str or date_str == "TBD":
        return "TBD"
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d").date()
        today = date.today()
        delta = (d - today).days
        if delta < 0:
            return "Past"
        elif delta == 0:
            return "Today"
        elif delta == 1:
            return "Tomorrow"
        # Use actual Mon–Fri calendar week boundaries
        monday_this_week = today - timedelta(days=today.weekday())
        friday_this_week = monday_this_week + timedelta(days=4)
        monday_next_week = monday_this_week + timedelta(days=7)
        friday_next_week = monday_next_week + timedelta(days=4)
        if d <= friday_this_week:
            return "This Week"
        elif d <= friday_next_week:
            return "Next Week"
        elif delta <= 30:
            return "Next 30 Days"
        else:
            return "Later"
    except Exception:
        return "TBD"


# ── Financial Summary ─────────────────────────────────────────────────────────

def fetch_financial_summary(api_key: str, code: str = None, date_str: str = None) -> list:
    """
    Fetch quarterly financial summary for a company or all companies on a date.
    Provide either code (e.g. "7203") or date_str (e.g. "2024-10-25").

    Returns list of summary dicts. Key fields:
      DiscDate, Code, CurPerType (1Q/2Q/3Q/4Q/FY),
      Sales, OP, NP, EPS,              ← actuals
      FSales, FOP, FNP, FEPS,          ← current year forecast
      NxFSales, NxFOP, NxFNP, NxFEPS,  ← next year forecast
      DivAnn, FDivAnn,                  ← actual + forecast annual dividend
    """
    if not api_key:
        return []
    if not code and not date_str:
        return []

    results = []
    pagination_key = None

    while True:
        params = {}
        if code:
            # J-Quants accepts 4 or 5 digit code
            params["code"] = code.replace(".T", "")
        if date_str:
            params["date"] = date_str
        if pagination_key:
            params["pagination_key"] = pagination_key

        try:
            resp = requests.get(
                f"{BASE}/fins/summary",
                headers=_headers(api_key),
                params=params,
                timeout=20,
            )
            if resp.status_code != 200:
                print(f"Financial summary error: HTTP {resp.status_code} — {resp.text[:100]}")
                break

            data = resp.json()
            batch = data.get("data", [])
            results.extend(batch)

            pagination_key = data.get("pagination_key")
            if not pagination_key or not batch:
                break

        except Exception as e:
            print(f"Financial summary fetch error: {e}")
            break

    # Sort by disclosure date ascending
    results.sort(key=lambda x: x.get("DiscDate", ""))
    return results


def format_summary_for_display(summaries: list) -> list:
    """
    Take raw financial summary records and return a cleaned list
    suitable for display — last 8 quarters, most recent first.
    """
    # Filter to main consolidated or standalone filings only
    # DocType contains 'FinancialStatements' or 'EarningsSummary'
    relevant = [
        s for s in summaries
        if s.get("CurPerType") in ("1Q", "2Q", "3Q", "4Q", "FY")
    ]
    # Most recent first
    relevant.sort(key=lambda x: x.get("DiscDate", ""), reverse=True)
    return relevant[:8]


def safe_num(val, divisor=1e9, decimals=1) -> str:
    """Format a large number in billions with ¥ prefix, or return '—'."""
    try:
        if val == "" or val is None:
            return "—"
        n = float(val) / divisor
        return f"¥{n:,.{decimals}f}B"
    except Exception:
        return "—"


def safe_pct(val, decimals=1) -> str:
    """Format a ratio as percentage, or return '—'."""
    try:
        if val == "" or val is None:
            return "—"
        return f"{float(val)*100:.{decimals}f}%"
    except Exception:
        return "—"


def guidance_direction(current: str, previous: str) -> str:
    """
    Compare two guidance figures and return direction emoji.
    Returns '▲ raised', '▼ cut', '— unchanged', or '' if unknown.
    """
    try:
        c = float(current)
        p = float(previous)
        if c > p * 1.005:
            return "▲ raised"
        elif c < p * 0.995:
            return "▼ cut"
        else:
            return "— unchanged"
    except Exception:
        return ""


# ── Price performance helpers ─────────────────────────────────────────────────

def get_performance_band(vs_topix_pct: float) -> dict:
    """
    Given a stock's return vs TOPIX (in percentage points), return
    display metadata: colour, label, emoji.
    """
    if vs_topix_pct is None:
        return {"color": "#9B8B7A", "bg": "#F0EDE8", "label": "N/A", "emoji": "●"}
    if vs_topix_pct >= 15:
        return {"color": "#1B5E20", "bg": "#C8E6C9", "label": f"+{vs_topix_pct:.1f}%", "emoji": "🟢"}
    elif vs_topix_pct >= 5:
        return {"color": "#2E7D32", "bg": "#DCEDC8", "label": f"+{vs_topix_pct:.1f}%", "emoji": "🟩"}
    elif vs_topix_pct >= -5:
        return {"color": "#6B4C00", "bg": "#FFF9C4", "label": f"{vs_topix_pct:+.1f}%", "emoji": "🟨"}
    elif vs_topix_pct >= -15:
        return {"color": "#E65100", "bg": "#FFE0B2", "label": f"{vs_topix_pct:.1f}%", "emoji": "🟧"}
    else:
        return {"color": "#B71C1C", "bg": "#FFCDD2", "label": f"{vs_topix_pct:.1f}%", "emoji": "🔴"}


def parse_jpx_earnings_excel(file_bytes: bytes, source_label: str = "") -> list:
    """
    Parse a JPX earnings announcement Excel file.

    Actual structure (confirmed from real files):
      Sheet name: 'List'
      Rows 1-4:  Title rows (skip)
      Row 5:     Headers
      Row 6+:    Data

    Columns (0-indexed):
      0: announcement date  (text: "2026-04-03 00:00:00")
      1: TSE code           (text: "2796")
      2: Company name JP
      3: Company name EN
      4: Fiscal year end    (text: "2026-03-31 00:00:00")
      5: Sector JP
      6: Sector EN
      7: Period type JP     ("本決算", "第１四半期", etc.)
    """
    try:
        import openpyxl
        import io
    except ImportError:
        print("openpyxl not installed")
        return []

    PERIOD_MAP = {
        "本決算": "Full Year",
        "第１四半期": "Q1", "第1四半期": "Q1",
        "第２四半期": "Q2", "第2四半期": "Q2",
        "第３四半期": "Q3", "第3四半期": "Q3",
        "第４四半期": "Q4", "第4四半期": "Q4",
        "中間": "Half Year", "半期": "Half Year",
    }

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        print(f"Excel open error ({source_label}): {e}")
        return []

    # Use sheet named 'List', or fall back to active sheet
    ws = wb["List"] if "List" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))

    results = []
    # Data starts at row index 5 (row 6 in Excel, after 4 title rows + 1 header row)
    for row in rows[5:]:
        if not row or (not row[0] and not row[1]):
            continue

        # Date: text like "2026-04-03 00:00:00" — take first 10 chars
        raw = str(row[0]).strip() if row[0] is not None else ""
        ann_date = raw[:10] if len(raw) >= 10 and "-" in raw[:10] else ""

        # Code: 4-digit string
        code_raw = str(row[1]).strip().replace(".0", "") if row[1] is not None else ""
        if not (code_raw.isdigit() and len(code_raw) == 4):
            continue  # skip header repeats and non-data rows

        # Company name: prefer English (col 3), fall back to Japanese (col 2)
        name = (str(row[3]).strip() if row[3] else "") or (str(row[2]).strip() if row[2] else "")

        # FY end date
        fy_raw = str(row[4]).strip() if row[4] is not None else ""
        fy_end = fy_raw[:10] if len(fy_raw) >= 10 and "-" in fy_raw[:10] else fy_raw

        # Sector (English, col 6)
        sector = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""

        # Period type (Japanese col 7 → English)
        period_raw = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ""
        period = PERIOD_MAP.get(period_raw, period_raw)

        results.append({
            "announcement_date": ann_date,
            "code":              code_raw,
            "name":              name,
            "sector":            sector,
            "period_type":       period,
            "fiscal_year_end":   fy_end,
            "source":            source_label,
        })

    return results


def fetch_jpx_excel_from_github(repo: str, branch: str = "main", token: str = None) -> list:
    """
    Download and parse all JPX Excel files committed to a GitHub repo
    under the path data/jpx_earnings/*.xlsx.

    repo:   e.g. "chernyeh/japan.news.digest"
    token:  GitHub personal access token (required for private repos)
    Returns combined list of earnings entries.
    """
    import requests as _req

    # Build request headers — include token if provided
    headers = {"Accept": "application/vnd.github.v3+json"}
    if token:
        headers["Authorization"] = f"token {token}"

    # Get the directory listing via GitHub API
    api_url = f"https://api.github.com/repos/{repo}/contents/data/jpx_earnings"
    try:
        resp = _req.get(api_url, timeout=10, headers=headers)
        if resp.status_code == 404:
            if not token:
                print("GitHub API 404 — repo may be private. Add GITHUB_TOKEN to Streamlit Secrets.")
            else:
                print(f"GitHub API 404 — check repo name and that token has repo access: {repo}")
            return []
        if resp.status_code != 200:
            print(f"GitHub API error: {resp.status_code} — {resp.text[:100]}")
            return []

        files = [f for f in resp.json() if f.get("name", "").endswith(".xlsx")]
        print(f"Found {len(files)} Excel file(s) in data/jpx_earnings/")
    except Exception as e:
        print(f"GitHub listing error: {e}")
        return []

    all_entries = []
    for f in files:
        fname = f["name"]
        try:
            # Use authenticated request for private repos
            dl_url = f["download_url"]
            file_resp = _req.get(dl_url, timeout=20, headers=headers)
            if file_resp.status_code == 200:
                entries = parse_jpx_earnings_excel(file_resp.content, source_label=fname)
                print(f"  {fname}: {len(entries)} entries")
                all_entries.extend(entries)
            else:
                print(f"  {fname}: HTTP {file_resp.status_code}")
        except Exception as e:
            print(f"  {fname}: download error {e}")

    # Sort by announcement date
    all_entries.sort(key=lambda x: x.get("announcement_date") or "9999")
    return all_entries


def fetch_market_data_batch(codes: list) -> dict:
    """
    Fetch market cap and 3M vs TOPIX performance for a list of TSE codes.
    Designed for batches of 20-500 stocks from the filtered earnings view.

    Returns dict: {code: {market_cap_b: float|None, vs_topix_3m: float|None}}
    market_cap_b = market cap in JPY billions
    vs_topix_3m  = 3M return minus TOPIX 3M return, in percentage points
    """
    try:
        import yfinance as yf
        import pandas as pd
        import requests as _req
    except ImportError:
        return {}

    if not codes:
        return {}

    results = {c: {"market_cap_b": None, "vs_topix_3m": None} for c in codes}

    # ── Step 1: Market cap via Yahoo Finance v7 quote API (batch, no auth) ──
    # Chunk into batches of 50 to avoid URL length limits
    tickers_str_list = [f"{c}.T" for c in codes]
    for i in range(0, len(tickers_str_list), 50):
        batch = tickers_str_list[i:i+50]
        symbols = ",".join(batch)
        url = f"https://query1.finance.yahoo.com/v7/finance/quote?symbols={symbols}&fields=marketCap,regularMarketPrice"
        try:
            resp = _req.get(url, timeout=15,
                           headers={"User-Agent": "Mozilla/5.0",
                                    "Accept": "application/json"})
            if resp.status_code == 200:
                data = resp.json()
                quotes = data.get("quoteResponse", {}).get("result", [])
                for q in quotes:
                    sym = q.get("symbol", "")
                    code = sym.replace(".T", "")
                    mc = q.get("marketCap")
                    if mc and mc > 0 and code in results:
                        results[code]["market_cap_b"] = round(mc / 1e9, 1)
        except Exception as e:
            print(f"Market cap batch error (chunk {i//50}): {e}")

    # ── Step 2: 3M vs TOPIX via yfinance price download ─────────────────────
    # First get TOPIX benchmark
    topix_3m = None
    try:
        topix_data = yf.download("1306.T", period="4mo", auto_adjust=True, progress=False)
        if not topix_data.empty:
            closes = topix_data["Close"].dropna()
            if len(closes) >= 2:
                cutoff = closes.index[-1] - pd.Timedelta(days=91)
                old = closes[closes.index <= cutoff]
                if len(old) > 0:
                    topix_3m = float((closes.iloc[-1] / old.iloc[-1] - 1) * 100)
    except Exception as e:
        print(f"TOPIX 3M error: {e}")

    # Batch download stock prices
    tickers = [f"{c}.T" for c in codes]
    try:
        if len(tickers) == 1:
            raw = yf.download(tickers[0], period="4mo", auto_adjust=True, progress=False)
            price_map = {tickers[0]: raw} if not raw.empty else {}
        else:
            raw = yf.download(tickers, period="4mo", auto_adjust=True,
                              progress=False, group_by="ticker")
            price_map = {}
            if not raw.empty:
                for t in tickers:
                    try:
                        df = raw[t].dropna(subset=["Close"])
                        if not df.empty:
                            price_map[t] = df
                    except Exception:
                        pass
    except Exception as e:
        print(f"Price batch error: {e}")
        price_map = {}

    # Compute 3M returns
    for code in codes:
        ticker = f"{code}.T"
        try:
            if ticker not in price_map:
                continue
            closes = price_map[ticker]["Close"].dropna()
            if len(closes) < 2:
                continue
            cutoff = closes.index[-1] - pd.Timedelta(days=91)
            old = closes[closes.index <= cutoff]
            if len(old) == 0:
                continue
            stock_3m = float((closes.iloc[-1] / old.iloc[-1] - 1) * 100)
            if topix_3m is not None:
                results[code]["vs_topix_3m"] = round(stock_3m - topix_3m, 1)
            else:
                results[code]["vs_topix_3m"] = round(stock_3m, 1)
        except Exception:
            pass

    return results


# Keep old name as alias for backwards compatibility
def fetch_3m_performance_batch(codes: list) -> dict:
    """Legacy wrapper — returns {code: vs_topix_pct}"""
    data = fetch_market_data_batch(codes)
    return {c: v.get("vs_topix_3m") for c, v in data.items()}




def load_mktcap_from_github(repo: str, token: str = None) -> dict:
    """
    Load market cap from mktcap_full.csv (primary) or jquants_prices_latest.csv (fallback).
    mktcap_full.csv is generated locally from archive prices × metadata shares.
    Returns dict: {code_4digit: market_cap_b_float}
    """
    import requests, csv, io

    BASE = f"https://raw.githubusercontent.com/{repo}/main/data"
    headers = {}
    if token:
        headers["Authorization"] = f"token {token}"

    # Try mktcap_full.csv first (generated locally, full coverage)
    for filename in ["mktcap_full.csv", "jquants_prices_latest.csv"]:
        url = f"{BASE}/{filename}"
        r = requests.get(url, headers=headers, timeout=15)
        if r.status_code != 200:
            continue

        mktcap = {}
        for row in csv.DictReader(io.StringIO(r.text)):
            code = str(row.get("Code", "")).strip().zfill(4)[:4]
            if not code or not code.isdigit():
                continue
            try:
                mc = row.get("MarketCapB", "")
                if mc and mc not in ("", "None", "null"):
                    mktcap[code] = float(mc)
            except (ValueError, TypeError):
                pass

        if mktcap:
            print(f"Market cap loaded: {len(mktcap)} companies (from {filename})")
            return mktcap

    print("Market cap: no data found")
    return {}


def load_3m_perf_from_github(repo: str, token: str = None) -> dict:
    """
    Load 3M vs TOPIX performance for all TSE stocks from the weekly cron CSV.
    Returns dict: {code_4digit: vs_topix_3m_float}
    """
    import requests, csv, io, base64

    headers = {"Accept": "application/vnd.github.v3.raw"}
    if token:
        headers["Authorization"] = f"token {token}"

    url = f"https://raw.githubusercontent.com/{repo}/main/data/jquants_3m_perf.csv"
    r = requests.get(url, headers=headers, timeout=15)
    if r.status_code == 404:
        return {}  # Not yet generated by cron
    if r.status_code != 200:
        print(f"3M perf fetch error: {r.status_code}")
        return {}

    raw = r.text
    reader = csv.DictReader(io.StringIO(raw))
    result = {}
    for row in reader:
        code = str(row.get("Code", "")).strip().zfill(4)[:4]
        if not code or not code.isdigit():
            continue
        try:
            val = row.get("VsTopix3M", "")
            if val not in ("", "None", "null"):
                result[code] = float(val)
        except (ValueError, TypeError):
            pass
    print(f"3M perf loaded: {len(result)} companies")
    return result


def load_perf_map_from_github(repo: str, token: str = None) -> tuple:
    """
    Load 3M/6M/12M vs TOPIX performance from the extended jquants_3m_perf.csv.
    Returns (perf_dict, topix_returns) where:
      perf_dict     = {code: {"vs3m": float|None, "vs6m": float|None, "vs12m": float|None}}
      topix_returns = {"3M": float|None, "6M": float|None, "12M": float|None}
    """
    import requests, csv, io

    headers = {"Accept": "application/vnd.github.v3.raw"}
    if token:
        headers["Authorization"] = f"token {token}"

    url = f"https://raw.githubusercontent.com/{repo}/main/data/jquants_3m_perf.csv"
    r = requests.get(url, headers=headers, timeout=15)
    if r.status_code != 200:
        print(f"Perf map fetch error: {r.status_code}")
        return {}, {"3M": None, "6M": None, "12M": None}

    def _f(v):
        if v not in ("", "None", "null", None):
            try:
                return float(v)
            except (ValueError, TypeError):
                pass
        return None

    perf_dict = {}
    topix_returns = {"3M": None, "6M": None, "12M": None}
    topix_set = False

    for row in csv.DictReader(io.StringIO(r.text)):
        code = str(row.get("Code", "")).strip().zfill(4)[:4]
        if not code or not code.isdigit():
            continue
        perf_dict[code] = {
            "vs3m":  _f(row.get("VsTopix3M")),
            "vs6m":  _f(row.get("VsTopix6M")),
            "vs12m": _f(row.get("VsTopix12M")),
        }
        if not topix_set:
            topix_returns["3M"]  = _f(row.get("TopixReturn3M"))
            topix_returns["6M"]  = _f(row.get("TopixReturn6M"))
            topix_returns["12M"] = _f(row.get("TopixReturn12M"))
            if any(v is not None for v in topix_returns.values()):
                topix_set = True

    print(f"Perf map loaded: {len(perf_dict)} companies")
    return perf_dict, topix_returns


def load_prices_from_github(repo: str, token: str = None) -> dict:
    """
    Load latest close prices from jquants_prices_latest.csv.
    Returns {code_4digit: close_price_float}
    """
    import requests, csv, io

    headers = {}
    if token:
        headers["Authorization"] = f"token {token}"

    url = f"https://raw.githubusercontent.com/{repo}/main/data/jquants_prices_latest.csv"
    r = requests.get(url, headers=headers, timeout=15)
    if r.status_code != 200:
        return {}

    prices = {}
    for row in csv.DictReader(io.StringIO(r.text)):
        code = str(row.get("Code", "")).strip().zfill(4)[:4]
        if not code or not code.isdigit():
            continue
        try:
            p = row.get("Close", "")
            if p not in ("", "None", "null"):
                prices[code] = float(p)
        except (ValueError, TypeError):
            pass

    print(f"Prices loaded: {len(prices)} companies")
    return prices


def load_earnings_cal_from_github(repo: str, token: str = None) -> list:
    """
    Load all JPX earnings calendar entries directly from GitHub repo.
    Reads all .xlsx files in data/jpx_earnings/ and parses them.
    Returns list of entry dicts (same format as parse_jpx_earnings_excel).
    """
    import requests, base64, io

    # For public repos no token needed; use plain headers to avoid rate-limit issues
    headers = {"Accept": "application/vnd.github.v3+json",
               "User-Agent": "japan-news-digest-app"}
    if token:
        headers["Authorization"] = f"token {token}"

    # List files in the folder
    url = f"https://api.github.com/repos/{repo}/contents/data/jpx_earnings"
    r = requests.get(url, headers=headers, timeout=15)
    if r.status_code == 404:
        return []
    if r.status_code != 200:
        print(f"Earnings cal folder fetch error: {r.status_code}")
        return []

    files = [f for f in r.json()
             if isinstance(f, dict) and f.get("name", "").lower().endswith(".xlsx")]
    if not files:
        return []

    all_entries = []
    for f in files:
        try:
            dl = requests.get(f["download_url"],
                              headers={"Authorization": f"token {token}"} if token else {},
                              timeout=30)
            if dl.status_code == 200:
                parsed = parse_jpx_earnings_excel(dl.content, source_label=f["name"])
                all_entries.extend(parsed)
        except Exception as e:
            print(f"Error loading {f['name']}: {e}")

    # Deduplicate by (code, announcement_date, period_type) — multiple Excel files can overlap
    seen = set()
    unique_entries = []
    for e in all_entries:
        key = (e.get("code"), e.get("announcement_date"), e.get("period_type"))
        if key not in seen:
            seen.add(key)
            unique_entries.append(e)
    all_entries = unique_entries

    all_entries.sort(key=lambda x: x.get("announcement_date") or "9999")
    print(f"Earnings calendar loaded: {len(all_entries)} entries from {len(files)} file(s)")
    return all_entries


def filter_upcoming(entries: list, days_ahead: int = 60) -> list:
    """Filter to entries with announcement dates within the next N days."""
    from datetime import date as _date, timedelta as _td
    today = _date.today()
    cutoff = today + _td(days=days_ahead)
    today_str   = today.strftime("%Y-%m-%d")
    cutoff_str  = cutoff.strftime("%Y-%m-%d")
    return [
        e for e in entries
        if e.get("announcement_date") and
           today_str <= e["announcement_date"] <= cutoff_str
    ]
